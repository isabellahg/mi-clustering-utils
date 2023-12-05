import glob
import matplotlib.pyplot as plt
from io import BytesIO
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import pandas as pd
import openpyxl
from kneebow.rotor import Rotor

def get_eps(distances):
    # Assuming distances is a sorted 1D numpy array of distances
    curve_xy = np.vstack((np.arange(len(distances)), distances)).T
    rotor = Rotor()
    rotor.fit_rotate(curve_xy)
    _, knee_y = rotor.get_elbow_point()
    return knee_y

def analyze_and_write_plot(sheet_name, book, newBook, file_path, file_path_new):
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    
    data_column = df.columns[0]
    scatter_data = []
    current_minpoints = 0
    data_start = 0

    for index, value in enumerate(df[data_column]):
        if pd.notnull(value) and 'minPoints' in str(value):
            scatter_data = []
            data_start = index
            current_minpoints = value
        
        elif pd.notnull(value) and 'Distance' not in str(value):
            scatter_data.append(value)

        elif pd.isnull(value) and scatter_data:
            print(scatter_data)

            scatter_data = np.sort(np.array(scatter_data))
            eps = get_eps(scatter_data)
            config = f'-E {eps:.2f} -M {current_minpoints} -D {euclidean}'
            plt.figure()
            plt.scatter(range(len(scatter_data)), scatter_data)
            plt.title(f'Scatter Plot for {current_minpoints}')
            img_data = BytesIO()
            plt.savefig(img_data, format='png')
            img_data.seek(0)
            img = Image(img_data)
            newBook[sheet_name].add_image(img, 'E' + str(data_start + 2))
            plt.close()
            scatter_data = [] 
        

    if scatter_data:
        plt.figure()
        plt.scatter(range(len(scatter_data)), scatter_data)
        plt.title(f'Scatter Plot for {current_minpoints}')
        
        img_data = BytesIO()
        plt.savefig(img_data, format='png')
        img_data.seek(0)
        
        img = Image(img_data)
        newBook[sheet_name].add_image(img, 'E' + str(data_start + 2))
        
        plt.close()
        
        scatter_data = [] 
    
file_path = 'reports'
result_path = 'charts'

excel_files = glob.glob(file_path + '/*.xlsx')

for file in excel_files:
    book = load_workbook(file)
    new_book = load_workbook(file)  
    sheet_names = book.sheetnames

    new_file_path = file.replace('.xlsx', '_generated.xlsx').replace(file_path, 'charts')

    for sheet in sheet_names:
        analyze_and_write_plot(sheet, book, new_book, file, new_file_path)

    new_book.save(new_file_path)